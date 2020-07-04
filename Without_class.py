'''A seller of luxury wrist watches has enlisted your help in performing some difficult tasks.
Combining any of lambda, map, filter or reduce, write a python program to:
The most expensive watch
A new inventory where the watches are sorted according to their price
A dictionary within which each key represents a watch brand and the value is the total
cost of that brand in our inventory (quantity * price)
A list containing all watches with quantity greater than 30
A new inventory with discount prices for items that are not getting sold out. In this new
inventory, if the quantity of a watch from the old inventory is greater than 30, reduce the
price by 50 EUR if the old price is greater than 500 EUR else reduce by 30 EUR'''

import openpyxl
from functools import reduce

wb = openpyxl.load_workbook('hm.xlsx')
sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column
watch = []
for i in range(4, rows + 1):
    for j in range(2, cols + 1, 2):
        cell = sheet.cell(row=i, column=j)
        if cell.value is None:
            continue
        watch.append(str(cell.value).strip())
id = []
item = []
quantity = []
unit_price = []
for i in range(0, len(watch), 4):
    id.append(int(watch[i]))
    item.append(watch[i + 1].strip())
    quantity.append(int(watch[i + 2]))
    unit_price.append(float(watch[i + 3].strip()))
inventory = []
for i, v in enumerate(id):
    inventory.append((id[i], item[i], quantity[i], unit_price[i]))
print(inventory)

print('---The most expensive watch---')
expensive_watch = reduce(max, unit_price)
print(f'The most expensive watch = {expensive_watch}')

print('---Sorted by price---')
s = sorted(unit_price, key=lambda x: x)
print(s)
print('--or--')
s = sorted(inventory, key=lambda x: x[3])
print(s)

print('---Dictionary---')
d = dict(map(lambda x: [x[1], round((x[2]) * (x[3]), 2)], inventory))
print(d)
print('--or--')

def make_dict(inventory):
    d2 = {}
    for i in inventory:
        watch_brand = i[1]
        total_cost = round((i[2] * i[3]), 2)
        d2.update({watch_brand: total_cost})
    return d2
print(make_dict(inventory))

print("---New inventory---")
def new_inventory(inventory):
    not_sold_out = []
    new_inv = []
    for i in inventory:
        if i[2] == 0:
            continue
        not_sold_out.append((i[0], i[1], i[2], i[3]))
    for i, v in enumerate(not_sold_out):
        not_sold_out[i] = list(not_sold_out[i])
        if not_sold_out[i][2] > 30:  # if quantity of watches > 30
           if not_sold_out[i][3] > 500:  # if price > 500
               not_sold_out[i][3] -= 50
           else:                 # if price <= 500
               not_sold_out[i][3] -= 30
        not_sold_out[i] = tuple(not_sold_out[i])
    inventory = not_sold_out
    for i, v in enumerate(inventory):
        print('(',inventory[i][1], '-', inventory[i][2],'items)', ' -> ', inventory[i][3],'eur')
        new_inv.append((inventory[i][0], inventory[i][1], inventory[i][2], inventory[i][3]))
    return new_inv

print(new_inventory(inventory))




'''
# id =[]
# for cell in sheet["B"]:
#     if cell.value is not None:
#         id.append(cell.value)
# print(id)
# item =[]
# for cell in sheet["D"]:
#     if cell.value is not None:
#         cell.value = cell.value.strip()
#         item.append(cell.value)
# quantity =[]
# for cell in sheet["F"]:
#     if cell.value is not None:
#         quantity.append(cell.value)
# unit_price =[]
# for cell in sheet["H"]:
#     if cell.value is not None:
#         unit_price.append(cell.value)
# inventory = []
# for i, v in enumerate(id):
#     if i == 0 or i == 1:
#         continue
#     inventory.append((id[i], item[i], quantity[i], float(unit_price[i])))
#

# # The most expensive watch
# print('---The most expensive watch---')
# res = reduce(max, map(lambda x: x[3], inventory))
# print(f"The most expensive watch -> {res}")
# price = []
# for i in inventory:
#     price.append(i[3])
# res2 = reduce(max, price)
# print(f"The most expensive watch -> {res2}")

# print('---Sorted by price---')
# # A new inventory where the watches are sorted according to their price
# s = sorted(inventory, key=lambda x: x[3])
# # print('sorted-price ->', s)
# for i, v in enumerate(s):
#     print(f'{s[i][1]} -> {s[i][3]}eur')
#
# print('---Dictionary---')
# # A dictionary within which each key represents a watch brand and the value is the total cost of that
# # brand in our inventory (quantity * price)
# d = dict(map(lambda x: [x[1], round(x[2]*x[3], 2)], inventory))
# print('dict ->', d)
#
# print('---All watches with quantity greater than 30---')
# # A list containing all watches with quantity greater than 30
# gt = list(filter(lambda x: x[2] > 30, inventory))
# # print('quantity > 30 ->', gt)
# for i, v in enumerate(gt):
#     print(f'{gt[i][1]} - {gt[i][2]}')
#
'''
# # A new inventory with discount prices for items that are not getting sold out.
# # In this new inventory, if the quantity of a watch from the old inventory is greater than 30,
# # reduce the price by 50 EUR if the old price is greater than 500 EUR else reduce by 30 EUR
#
# # disc = list(filter(lambda x: x[2] > 0, sh))
# # m = list(map(lambda x: x[2] > 30 and x[3]-50 or x[3] > 500 and x[3]-30 or x[3], disc))
# # # print('m', m)
# # # print(disc)
# # for i, v in enumerate(disc):
# #     disc[i] = list(disc[i])
# #     disc[i][3] = m[i]
# #     disc[i] = tuple(disc[i])
# # sh = disc
# # print('new -> ', sh)
'''
# print("---New inventory---")
# not_sold_out = list(filter(lambda x: x[2] > 0, inventory))
# for i, v in enumerate(not_sold_out):
#     not_sold_out[i] = list(not_sold_out[i])
#     if not_sold_out[i][2] > 30:  # if quantity of watches > 30
#         if not_sold_out[i][3] > 500:  # if price > 500
#             not_sold_out[i][3] -= 50
#         else:                 # if price <= 500
#             not_sold_out[i][3] -= 30
#     not_sold_out[i] = tuple(not_sold_out[i])
# inventory = not_sold_out
# for i, v in enumerate(inventory):
#     print('(',inventory[i][1], '-', inventory[i][2],'items)', ' -> ', inventory[i][3],'eur')
'''
